import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
# # 实例化
# wb = Workbook()
# # 激活 worksheet,这个函数使用_active_sheet_index这个属性，默认设置的值是0，除非你指定一个值，否则总是获取到第一个worksheet。
# ws = wb.active
#
# # 单元格直接赋值
# ws['A1'] = 42
# ws['A2'] = '=sum(B2:C2)'
# wb.save('练习3.xlsx')
#
# print(ws['A2'].value)
# # wb2 = load_workbook('练习2.xlsx')
wb = Workbook()
ws = wb.active

for r in dataframe_to_rows(df, index=True, header=True):
    ws.append(r)

for cell in ws['A'] + ws[1]:
    cell.style = 'Pandas'

wb.save("pandas_openpyxl.xlsx")